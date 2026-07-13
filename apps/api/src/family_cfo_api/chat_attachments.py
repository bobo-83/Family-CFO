"""M85: bounded, grounded previews of data files attached to advisor chat.

A user can attach a CSV, spreadsheet (.xlsx), or plain-text file to a chat
message. The model cannot ingest the raw file, so the server extracts a
BOUNDED structured summary — column headers, row count, per-column numeric
stats, and the first rows — that joins the prompt as grounded context. The
numbers in the summary come from the user's own file, so they are legitimate
to echo (added to the guardrail's known values by the caller). Nothing is
written to the household's records; a chat attachment is read-only context,
distinct from the M7 imports pipeline.

Defensive by design: a malformed or unreadable file yields a best-effort note,
never an exception.
"""

from __future__ import annotations

import csv
import io

# Bounds — a chat attachment must never balloon the prompt or the memory used
# to parse it.
_MAX_ROWS_SCANNED = 5_000
_MAX_ROWS_SHOWN = 15
_MAX_COLS = 40
_MAX_PREVIEW_CHARS = 4_000
_MAX_TEXT_LINES = 40


def _is_number(value: str) -> float | None:
    text = value.strip().replace(",", "").replace("$", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _summarize_columns(header: list[str], rows: list[list[str]]) -> list[str]:
    lines: list[str] = []
    for col_index, name in enumerate(header[:_MAX_COLS]):
        numbers: list[float] = []
        non_empty = 0
        for row in rows:
            if col_index >= len(row):
                continue
            cell = row[col_index]
            if cell.strip():
                non_empty += 1
            parsed = _is_number(cell)
            if parsed is not None:
                numbers.append(parsed)
        label = name.strip() or f"column {col_index + 1}"
        if numbers and len(numbers) >= max(1, non_empty // 2):
            total = sum(numbers)
            lines.append(
                f"- {label}: numeric, {len(numbers)} values, "
                f"min {min(numbers):,.2f}, max {max(numbers):,.2f}, sum {total:,.2f}"
            )
        else:
            lines.append(f"- {label}: text, {non_empty} non-empty values")
    return lines


def _render_table_preview(name: str, header: list[str], rows: list[list[str]], truncated: bool) -> str:
    parts = [f'Attached data file "{name}" — {len(rows)} data rows' + (" (truncated)" if truncated else "") + "."]
    parts.append(f"Columns ({min(len(header), _MAX_COLS)}): {', '.join(h.strip() or '?' for h in header[:_MAX_COLS])}")
    parts.append("Per-column summary:")
    parts.extend(_summarize_columns(header, rows))
    shown = rows[:_MAX_ROWS_SHOWN]
    if shown:
        parts.append(f"First {len(shown)} rows:")
        for row in shown:
            parts.append("  " + " | ".join(cell.strip() for cell in row[:_MAX_COLS]))
    return "\n".join(parts)[:_MAX_PREVIEW_CHARS]


def _preview_csv(name: str, raw: bytes) -> str:
    text = raw.decode("utf-8", errors="replace")
    try:
        dialect = csv.Sniffer().sniff(text[:4096], delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    reader = csv.reader(io.StringIO(text), dialect)
    all_rows = []
    for i, row in enumerate(reader):
        if i > _MAX_ROWS_SCANNED:
            break
        all_rows.append(row)
    if not all_rows:
        return f'Attached data file "{name}" appears to be empty.'
    header = all_rows[0]
    rows = all_rows[1:]
    truncated = len(rows) >= _MAX_ROWS_SCANNED
    return _render_table_preview(name, header, rows, truncated)


def _preview_xlsx(name: str, raw: bytes) -> str:
    try:
        import openpyxl
    except ImportError:  # pragma: no cover - dependency is declared
        return f'Attached spreadsheet "{name}" could not be read (no spreadsheet engine).'
    try:
        workbook = openpyxl.load_workbook(
            io.BytesIO(raw), read_only=True, data_only=True
        )
    except Exception:
        return f'Attached spreadsheet "{name}" could not be opened — it may be corrupt.'
    try:
        sheet = workbook[workbook.sheetnames[0]]
        all_rows: list[list[str]] = []
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if i > _MAX_ROWS_SCANNED:
                break
            all_rows.append(["" if cell is None else str(cell) for cell in row])
    finally:
        workbook.close()
    if not all_rows:
        return f'Attached spreadsheet "{name}" (sheet "{sheet.title}") is empty.'
    header = all_rows[0]
    rows = all_rows[1:]
    truncated = len(rows) >= _MAX_ROWS_SCANNED
    label = f'{name}" (sheet "{sheet.title}'
    return _render_table_preview(label, header, rows, truncated)


def _preview_text(name: str, raw: bytes) -> str:
    text = raw.decode("utf-8", errors="replace")
    lines = text.splitlines()
    head = "\n".join(lines[:_MAX_TEXT_LINES])
    note = (
        f'Attached text file "{name}" — {len(lines)} lines, {len(text)} characters. '
        f"First {min(len(lines), _MAX_TEXT_LINES)} lines:\n{head}"
    )
    return note[:_MAX_PREVIEW_CHARS]


def build_data_file_preview(name: str, raw: bytes) -> str:
    """A bounded, grounded text summary of the attached file (never raises)."""
    lowered = name.lower()
    try:
        if lowered.endswith((".xlsx", ".xlsm")):
            return _preview_xlsx(name, raw)
        if lowered.endswith((".csv", ".tsv")):
            return _preview_csv(name, raw)
        # A comma/tab-delimited body without a .csv name still reads as a table.
        sample = raw[:2048].decode("utf-8", errors="replace")
        if lowered.endswith((".txt", ".md", ".log")) or ("," not in sample and "\t" not in sample):
            return _preview_text(name, raw)
        return _preview_csv(name, raw)
    except Exception:  # pragma: no cover - final safety net
        return f'Attached file "{name}" could not be summarized.'
